from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q, Count, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from django.core.paginator import Paginator
from django.http import HttpResponse
import json
from .models import Complaint, Category, ComplaintStatusHistory
from .forms import ComplaintForm
from accounts.models import AuditLog
from accounts.views import log_audit, get_client_ip
from ml_spam.ml_models import spam_ml


def home(request):
    categories = Category.objects.all()
    total = Complaint.objects.filter(is_spam=False).count()
    resolved = Complaint.objects.filter(status='resolved', is_spam=False).count()
    
    return render(request, 'complaints/landing.html', {
        'categories': categories,
        'total_complaints': total,
        'resolved_complaints': resolved,
    })


@login_required
def submit_complaint(request):
    if request.method == 'POST':
        form = ComplaintForm(request.POST, request.FILES)
        if form.is_valid():
            complaint = form.save(commit=False)
            complaint.submitted_by = request.user
            
            text = f"{complaint.description} {complaint.location}"
            is_spam, confidence = spam_ml.predict(text)
            
            complaint.is_spam = is_spam
            complaint.spam_confidence = confidence
            complaint.save()
            
            ComplaintStatusHistory.objects.create(
                complaint=complaint,
                status='pending',
                changed_by=request.user,
                notes='Complaint submitted'
            )
            
            log_audit(request.user, 'create_complaint', f'Created complaint {complaint.tracking_code}', request)
            
            if is_spam:
                messages.warning(request, f'⚠️ Complaint flagged for review. Code: {complaint.tracking_code}')
            else:
                messages.success(request, f'✅ Complaint submitted! Code: {complaint.tracking_code}')
            
            return redirect('complaint_success', tracking_code=complaint.tracking_code)
    else:
        form = ComplaintForm()
    
    return render(request, 'complaints/submit_complaint.html', {'form': form})


def complaint_success(request, tracking_code):
    complaint = get_object_or_404(Complaint, tracking_code=tracking_code)
    return render(request, 'complaints/complaint_success.html', {'complaint': complaint})


def track_complaint(request):
    """Track complaint by tracking code (works for both POST and GET)"""
    
    # Handle GET request with ?code= parameter
    if request.method == 'GET':
        code = request.GET.get('code', '').upper().strip()
        if code:
            try:
                complaint = Complaint.objects.get(tracking_code=code)
                # Check if user has permission to view
                if request.user.is_authenticated and (complaint.submitted_by == request.user or request.user.is_admin()):
                    return render(request, 'complaints/track_result.html', {'complaint': complaint})
                elif request.user.is_authenticated:
                    messages.error(request, 'You can only view your own complaints.')
                else:
                    messages.error(request, 'Please login to view complaint details.')
            except Complaint.DoesNotExist:
                messages.error(request, 'No complaint found with that code.')
    
    # Handle POST request
    if request.method == 'POST':
        code = request.POST.get('tracking_code', '').upper().strip()
        if code:
            try:
                complaint = Complaint.objects.get(tracking_code=code)
                if request.user.is_authenticated and (complaint.submitted_by == request.user or request.user.is_admin()):
                    return render(request, 'complaints/track_result.html', {'complaint': complaint})
                elif request.user.is_authenticated:
                    messages.error(request, 'You can only track your own complaints.')
                else:
                    messages.error(request, 'Please login to track complaints.')
            except Complaint.DoesNotExist:
                messages.error(request, 'No complaint found with that tracking code.')
    
    return render(request, 'complaints/track_complaint.html')


@login_required
def my_complaints(request):
    complaints = Complaint.objects.filter(submitted_by=request.user).order_by('-created_at')
    return render(request, 'complaints/my_complaints.html', {'complaints': complaints})


@login_required
@user_passes_test(lambda u: u.is_admin())
def admin_dashboard(request):
    complaints = Complaint.objects.all().order_by('-created_at')
    
    # Filters
    date_range = request.GET.get('date_range', 'all')
    if date_range == 'today':
        complaints = complaints.filter(created_at__date=timezone.now().date())
    elif date_range == 'week':
        complaints = complaints.filter(created_at__gte=timezone.now() - timedelta(days=7))
    elif date_range == 'month':
        complaints = complaints.filter(created_at__gte=timezone.now() - timedelta(days=30))
    
    search = request.GET.get('search', '')
    if search:
        complaints = complaints.filter(
            Q(tracking_code__icontains=search) |
            Q(description__icontains=search) |
            Q(location__icontains=search) |
            Q(submitted_by__first_name__icontains=search) |
            Q(submitted_by__last_name__icontains=search)
        )
    
    # Stats
    total = complaints.count()
    stats = {
        'pending': complaints.filter(status='pending').count(),
        'under_review': complaints.filter(status='under_review').count(),
        'in_progress': complaints.filter(status='in_progress').count(),
        'resolved': complaints.filter(status='resolved').count(),
        'rejected': complaints.filter(status='rejected').count(),
        'spam': complaints.filter(is_spam=True).count(),
    }
    
    # Category data for chart
    category_data = []
    for cat in Category.objects.all():
        count = complaints.filter(category=cat).count()
        if count > 0:
            category_data.append({'name': cat.name, 'count': count})
    
    # Monthly data
    monthly_data = []
    for i in range(5, -1, -1):
        date = timezone.now().date() - timedelta(days=30*i)
        count = complaints.filter(
            created_at__year=date.year,
            created_at__month=date.month
        ).count()
        monthly_data.append({
            'month': date.strftime('%b'),
            'count': count
        })
    
    # Pending complaints count for dashboard
    pending_complaints_count = Complaint.objects.filter(
        is_spam=False,
        reviewed_by_admin=True
    ).exclude(status__in=['resolved', 'rejected']).count()
    
    context = {
        'complaints': complaints,
        'stats': stats,
        'total': total,
        'category_data': json.dumps(category_data),
        'monthly_data': json.dumps(monthly_data),
        'search': search,
        'date_range': date_range,
        'pending_complaints_count': pending_complaints_count,
    }
    return render(request, 'complaints/admin_dashboard.html', context)


@login_required
@user_passes_test(lambda u: u.is_admin())
def update_status(request, complaint_id):
    complaint = get_object_or_404(Complaint, id=complaint_id)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        notes = request.POST.get('notes', '')
        
        old_status = complaint.status
        complaint.status = new_status
        complaint.status_updated_by = request.user
        complaint.save()
        
        ComplaintStatusHistory.objects.create(
            complaint=complaint,
            status=new_status,
            changed_by=request.user,
            notes=notes
        )
        
        log_audit(request.user, 'update_status', f'Changed {complaint.tracking_code} from {old_status} to {new_status}', request)
        messages.success(request, f'Status updated to {new_status}')
    
    return redirect('admin_dashboard')


@login_required
@user_passes_test(lambda u: u.is_admin())
def manage_categories(request):
    categories = Category.objects.all()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            name = request.POST.get('name')
            desc = request.POST.get('description')
            Category.objects.create(name=name, description=desc)
            messages.success(request, f'Category "{name}" added')
            log_audit(request.user, 'create_complaint', f'Added category {name}', request)
        
        elif action == 'edit':
            cat_id = request.POST.get('category_id')
            cat = get_object_or_404(Category, id=cat_id)
            cat.name = request.POST.get('name')
            cat.description = request.POST.get('description')
            cat.save()
            messages.success(request, 'Category updated')
        
        elif action == 'delete':
            cat_id = request.POST.get('category_id')
            Category.objects.filter(id=cat_id).delete()
            messages.success(request, 'Category deleted')
        
        return redirect('manage_categories')
    
    return render(request, 'complaints/manage_categories.html', {'categories': categories})


@login_required
@user_passes_test(lambda u: u.is_admin())
def review_spam(request):
    spam_complaints = Complaint.objects.filter(is_spam=True, reviewed_by_admin=False).order_by('-spam_confidence')
    
    if request.method == 'POST':
        complaint_id = request.POST.get('complaint_id')
        action = request.POST.get('action')
        complaint = get_object_or_404(Complaint, id=complaint_id)
        
        if action == 'approve':
            complaint.is_spam = False
            complaint.reviewed_by_admin = True
            complaint.save()
            messages.success(request, f'Complaint {complaint.tracking_code} marked as legitimate')
            log_audit(request.user, 'review_spam', f'Approved {complaint.tracking_code}', request)
        
        elif action == 'delete':
            complaint.is_spam = True
            complaint.reviewed_by_admin = True
            complaint.status = 'rejected'
            complaint.save()
            messages.success(request, f'Spam complaint {complaint.tracking_code} rejected')
            log_audit(request.user, 'review_spam', f'Rejected {complaint.tracking_code}', request)
        
        return redirect('review_spam')
    
    return render(request, 'complaints/review_spam.html', {'complaints': spam_complaints})


@login_required
@user_passes_test(lambda u: u.is_captain())
def captain_dashboard(request):
    return render(request, 'complaints/captain_dashboard.html')


# ==================== RESOLVED & PENDING VIEWS ====================

@login_required
@user_passes_test(lambda u: u.is_admin())
def resolved_complaints(request):
    """View all resolved complaints with search, filters, and pagination"""
    complaints = Complaint.objects.filter(
        status='resolved',
        is_spam=False
    ).order_by('-resolved_at')
    
    # Search
    search = request.GET.get('search', '')
    if search:
        complaints = complaints.filter(
            Q(tracking_code__icontains=search) |
            Q(description__icontains=search) |
            Q(location__icontains=search) |
            Q(submitted_by__first_name__icontains=search) |
            Q(submitted_by__last_name__icontains=search) |
            Q(submitted_by__email__icontains=search)
        )
    
    # Date filter
    date_range = request.GET.get('date_range', 'all')
    today = timezone.now().date()
    if date_range == 'today':
        complaints = complaints.filter(resolved_at__date=today)
    elif date_range == 'week':
        start_date = today - timedelta(days=7)
        complaints = complaints.filter(resolved_at__date__gte=start_date)
    elif date_range == 'month':
        start_date = today - timedelta(days=30)
        complaints = complaints.filter(resolved_at__date__gte=start_date)
    elif date_range == 'year':
        start_date = today - timedelta(days=365)
        complaints = complaints.filter(resolved_at__date__gte=start_date)
    
    # Stats
    total_resolved = complaints.count()
    resolved_this_month = complaints.filter(resolved_at__month=today.month).count()
    
    # Average resolution time
    avg_seconds = 0
    for complaint in complaints:
        if complaint.resolved_at and complaint.created_at:
            diff = complaint.resolved_at - complaint.created_at
            avg_seconds += diff.total_seconds()
    
    if complaints.count() > 0:
        avg_seconds = avg_seconds / complaints.count()
        avg_hours = avg_seconds / 3600
        avg_days = avg_hours / 24
    else:
        avg_hours = 0
        avg_days = 0
    
    # Resolution rate
    total_complaints = Complaint.objects.filter(is_spam=False).count()
    if total_complaints > 0:
        resolution_rate = (total_resolved / total_complaints) * 100
    else:
        resolution_rate = 0
    
    # Pagination
    paginator = Paginator(complaints, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'complaints': page_obj,
        'total_resolved': total_resolved,
        'resolved_this_month': resolved_this_month,
        'avg_resolution_hours': round(avg_hours, 1),
        'avg_resolution_days': round(avg_days, 1),
        'resolution_rate': round(resolution_rate, 1),
        'search': search,
        'date_range': date_range,
    }
    return render(request, 'complaints/resolved_complaints.html', context)


@login_required
@user_passes_test(lambda u: u.is_admin())
def pending_complaints(request):
    """View all active/pending complaints with search, filters, and pagination"""
    complaints = Complaint.objects.filter(
        is_spam=False,
        reviewed_by_admin=True
    ).exclude(
        status__in=['resolved', 'rejected']
    ).order_by('-created_at')
    
    # Search
    search = request.GET.get('search', '')
    if search:
        complaints = complaints.filter(
            Q(tracking_code__icontains=search) |
            Q(description__icontains=search) |
            Q(location__icontains=search) |
            Q(submitted_by__first_name__icontains=search) |
            Q(submitted_by__last_name__icontains=search) |
            Q(submitted_by__email__icontains=search)
        )
    
    # Status filter
    status_filter = request.GET.get('status_filter', 'all')
    if status_filter != 'all':
        complaints = complaints.filter(status=status_filter)
    
    # Date filter
    date_range = request.GET.get('date_range', 'all')
    today = timezone.now().date()
    if date_range == 'today':
        complaints = complaints.filter(created_at__date=today)
    elif date_range == 'week':
        start_date = today - timedelta(days=7)
        complaints = complaints.filter(created_at__date__gte=start_date)
    elif date_range == 'month':
        start_date = today - timedelta(days=30)
        complaints = complaints.filter(created_at__date__gte=start_date)
    
    # Stats
    total_active = complaints.count()
    pending_count = complaints.filter(status='pending').count()
    under_review_count = complaints.filter(status='under_review').count()
    in_progress_count = complaints.filter(status='in_progress').count()
    
    # Pagination
    paginator = Paginator(complaints, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'complaints': page_obj,
        'total_active': total_active,
        'pending_count': pending_count,
        'under_review_count': under_review_count,
        'in_progress_count': in_progress_count,
        'search': search,
        'status_filter': status_filter,
        'date_range': date_range,
    }
    return render(request, 'complaints/pending_complaints.html', context)


@login_required
@user_passes_test(lambda u: u.is_admin())
def flag_as_spam(request, complaint_id):
    """Manually flag a complaint as spam"""
    complaint = get_object_or_404(Complaint, id=complaint_id)
    
    if request.method == 'POST':
        complaint.is_spam = True
        complaint.spam_confidence = 1.0
        complaint.reviewed_by_admin = True
        complaint.status = 'rejected'
        complaint.save()
        
        # Add to status history
        ComplaintStatusHistory.objects.create(
            complaint=complaint,
            status='rejected',
            changed_by=request.user,
            notes='Manually flagged as spam by admin'
        )
        
        log_audit(request.user, 'update_status', 
                 f'Manually flagged {complaint.tracking_code} as spam', request)
        messages.success(request, f'Complaint {complaint.tracking_code} has been flagged as spam.')
    
    return redirect(request.META.get('HTTP_REFERER', 'admin_dashboard'))


@login_required
@user_passes_test(lambda u: u.is_admin())
def mark_as_not_spam(request, complaint_id):
    """Mark a flagged complaint as not spam (false positive)"""
    complaint = get_object_or_404(Complaint, id=complaint_id)
    
    if request.method == 'POST':
        complaint.is_spam = False
        complaint.spam_confidence = 0.0
        complaint.reviewed_by_admin = True
        complaint.save()
        
        log_audit(request.user, 'update_status', 
                 f'Marked {complaint.tracking_code} as not spam', request)
        messages.success(request, f'Complaint {complaint.tracking_code} marked as legitimate.')
    
    return redirect(request.META.get('HTTP_REFERER', 'review_spam'))


# ==================== EXPORT FUNCTIONS ====================

@login_required
@user_passes_test(lambda u: u.is_admin())
def export_complaints_excel(request):
    """Export complaints to Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Get complaints based on filters
    complaints = Complaint.objects.all().order_by('-created_at')
    
    # Apply filters if any
    status_filter = request.GET.get('status', '')
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    
    date_range = request.GET.get('date_range', 'all')
    today = timezone.now().date()
    if date_range == 'today':
        complaints = complaints.filter(created_at__date=today)
    elif date_range == 'week':
        start_date = today - timedelta(days=7)
        complaints = complaints.filter(created_at__date__gte=start_date)
    elif date_range == 'month':
        start_date = today - timedelta(days=30)
        complaints = complaints.filter(created_at__date__gte=start_date)
    
    search = request.GET.get('search', '')
    if search:
        complaints = complaints.filter(
            Q(tracking_code__icontains=search) |
            Q(description__icontains=search) |
            Q(location__icontains=search) |
            Q(submitted_by__first_name__icontains=search) |
            Q(submitted_by__last_name__icontains=search)
        )
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Complaints"
    
    # Headers
    headers = ['Tracking Code', 'Complainant', 'Email', 'Category', 'Description', 
               'Location', 'Status', 'Is Spam', 'Created Date', 'Resolved Date']
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4E73DF", end_color="4E73DF", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row, complaint in enumerate(complaints, 2):
        ws.cell(row=row, column=1, value=complaint.tracking_code)
        ws.cell(row=row, column=2, value=f"{complaint.submitted_by.first_name} {complaint.submitted_by.last_name}")
        ws.cell(row=row, column=3, value=complaint.submitted_by.email)
        ws.cell(row=row, column=4, value=complaint.category.name if complaint.category else 'Uncategorized')
        ws.cell(row=row, column=5, value=complaint.description)
        ws.cell(row=row, column=6, value=complaint.location)
        ws.cell(row=row, column=7, value=complaint.get_status_display())
        ws.cell(row=row, column=8, value='Yes' if complaint.is_spam else 'No')
        ws.cell(row=row, column=9, value=complaint.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=10, value=complaint.resolved_at.strftime('%Y-%m-%d %H:%M') if complaint.resolved_at else '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"complaints_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.is_admin())
def export_complaints_csv(request):
    """Export complaints to CSV"""
    import csv
    
    # Get complaints based on filters
    complaints = Complaint.objects.all().order_by('-created_at')
    
    # Apply filters if any
    status_filter = request.GET.get('status', '')
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    
    date_range = request.GET.get('date_range', 'all')
    today = timezone.now().date()
    if date_range == 'today':
        complaints = complaints.filter(created_at__date=today)
    elif date_range == 'week':
        start_date = today - timedelta(days=7)
        complaints = complaints.filter(created_at__date__gte=start_date)
    elif date_range == 'month':
        start_date = today - timedelta(days=30)
        complaints = complaints.filter(created_at__date__gte=start_date)
    
    search = request.GET.get('search', '')
    if search:
        complaints = complaints.filter(
            Q(tracking_code__icontains=search) |
            Q(description__icontains=search) |
            Q(location__icontains=search) |
            Q(submitted_by__first_name__icontains=search) |
            Q(submitted_by__last_name__icontains=search)
        )
    
    # Create response
    response = HttpResponse(content_type='text/csv')
    filename = f"complaints_{timezone.now().strftime('%Y%m%d_%H%M')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow(['Tracking Code', 'Complainant', 'Email', 'Category', 'Description', 
                     'Location', 'Status', 'Is Spam', 'Created Date', 'Resolved Date'])
    
    # Write data
    for complaint in complaints:
        writer.writerow([
            complaint.tracking_code,
            f"{complaint.submitted_by.first_name} {complaint.submitted_by.last_name}",
            complaint.submitted_by.email,
            complaint.category.name if complaint.category else 'Uncategorized',
            complaint.description,
            complaint.location,
            complaint.get_status_display(),
            'Yes' if complaint.is_spam else 'No',
            complaint.created_at.strftime('%Y-%m-%d %H:%M'),
            complaint.resolved_at.strftime('%Y-%m-%d %H:%M') if complaint.resolved_at else ''
        ])
    
    return response


@login_required
@user_passes_test(lambda u: u.is_admin())
def export_complaints_pdf(request):
    """Export complaints to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    # Get complaints based on filters
    complaints = Complaint.objects.all().order_by('-created_at')
    
    # Apply filters if any
    status_filter = request.GET.get('status', '')
    if status_filter:
        complaints = complaints.filter(status=status_filter)
    
    date_range = request.GET.get('date_range', 'all')
    today = timezone.now().date()
    if date_range == 'today':
        complaints = complaints.filter(created_at__date=today)
    elif date_range == 'week':
        start_date = today - timedelta(days=7)
        complaints = complaints.filter(created_at__date__gte=start_date)
    elif date_range == 'month':
        start_date = today - timedelta(days=30)
        complaints = complaints.filter(created_at__date__gte=start_date)
    
    search = request.GET.get('search', '')
    if search:
        complaints = complaints.filter(
            Q(tracking_code__icontains=search) |
            Q(description__icontains=search) |
            Q(location__icontains=search) |
            Q(submitted_by__first_name__icontains=search) |
            Q(submitted_by__last_name__icontains=search)
        )
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    filename = f"complaints_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1
    
    # Title
    elements.append(Paragraph("Barangay Complaint System Report", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Date
    date_style = styles['Normal']
    date_style.alignment = 1
    elements.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y %I:%M %p')}", date_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Statistics
    elements.append(Paragraph("Summary Statistics", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))
    
    stats_data = [
        ['Metric', 'Count'],
        ['Total Complaints', str(complaints.count())],
        ['Pending', str(complaints.filter(status='pending').count())],
        ['Under Review', str(complaints.filter(status='under_review').count())],
        ['In Progress', str(complaints.filter(status='in_progress').count())],
        ['Resolved', str(complaints.filter(status='resolved').count())],
        ['Rejected', str(complaints.filter(status='rejected').count())],
        ['Flagged as Spam', str(complaints.filter(is_spam=True).count())],
    ]
    
    stats_table = Table(stats_data, colWidths=[2*inch, 1.5*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4E73DF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Complaints Table
    elements.append(Paragraph("Complaint Details", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))
    
    table_data = [['Tracking', 'Complainant', 'Category', 'Status', 'Date']]
    
    for complaint in complaints[:20]:  # Limit to 20 for PDF
        table_data.append([
            complaint.tracking_code,
            f"{complaint.submitted_by.first_name} {complaint.submitted_by.last_name}",
            complaint.category.name if complaint.category else 'Uncategorized',
            complaint.get_status_display(),
            complaint.created_at.strftime('%Y-%m-%d')
        ])
    
    complaint_table = Table(table_data, colWidths=[1.2*inch, 1.5*inch, 1.2*inch, 1*inch, 1.2*inch])
    complaint_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4E73DF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(complaint_table)
    
    doc.build(elements)
    return response